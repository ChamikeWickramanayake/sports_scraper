"""
Excel Export Module
Handles all interactions with local Excel files
"""
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from config.config import EXCEL_FILE, EXCEL_COLUMNS
from utils.logger import logger

# Control characters openpyxl refuses to write (IllegalCharacterError)
_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize(value):
    """Strip characters that openpyxl can't write to a cell."""
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS.sub("", value)
    return value


class ExcelExporter:
    """Exports events to a local Excel file"""
    
    def __init__(self, file_path=None):
        """
        Initialize Excel exporter
        
        Args:
            file_path (Path): Path to Excel file (default: config.EXCEL_FILE)
        """
        self.file_path = file_path or EXCEL_FILE
        self.columns = EXCEL_COLUMNS + ["Teams", "Source", "Timestamp"]
        self.workbook = None
        self.sheet = None
    
    def _load_or_create_workbook(self):
        """Load existing workbook or create new one"""
        try:
            if self.file_path.exists():
                logger.info(f"Loading existing Excel file: {self.file_path}")
                self.workbook = load_workbook(self.file_path)
                self.sheet = self.workbook.active
            else:
                logger.info(f"Creating new Excel file: {self.file_path}")
                self.workbook = Workbook()
                self.sheet = self.workbook.active
                self.sheet.title = "Sports Events"
                self._create_header_row()
        except Exception as e:
            logger.error(f"Failed to load/create workbook: {e}")
            raise
    
    def _create_header_row(self):
        """Create header row with formatting"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, column_name in enumerate(self.columns, start=1):
            cell = self.sheet.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        column_widths = {
            "Sport": 15,
            "Event": 30,
            "Broadcasting Partner": 20,
            "Event Date": 20,
            "Location": 20,
            "Teams": 30,
            "Source": 20,
            "Timestamp": 20,
        }
        for col_idx, column_name in enumerate(self.columns, start=1):
            self.sheet.column_dimensions[get_column_letter(col_idx)].width = column_widths.get(column_name, 15)
    
    def _get_existing_events(self):
        """Get set of existing events to avoid duplicates"""
        existing_events = set()
        
        if self.sheet.max_row > 1:  # Skip header row
            for row_idx in range(2, self.sheet.max_row + 1):
                row_data = []
                for col_idx in range(1, 6):  # first 5 columns; key uses Sport, Event, Event Date, Location
                    cell = self.sheet.cell(row=row_idx, column=col_idx)
                    row_data.append(str(cell.value) if cell.value else "")
                
                if row_data[0]:  # If Sport is not empty
                    event_key = (row_data[0], row_data[1], row_data[3], row_data[4])
                    existing_events.add(event_key)
        
        return existing_events
    
    def append_events(self, events):
        """
        Append events to Excel file, avoiding duplicates
        
        Args:
            events (list): List of event dictionaries
                {
                    "sport": str,
                    "event_name": str,
                    "broadcaster": str,
                    "event_date": str,
                    "location": str,
                    "teams": str,
                    "source": str,
                }
        """
        try:
            self._load_or_create_workbook()
            
            existing_events = self._get_existing_events()
            new_count = 0
            
            for event in events:
                event_key = (
                    event.get("sport", ""),
                    event.get("event_name", ""),
                    event.get("event_date", ""),
                    event.get("location", ""),
                )
                
                # Skip if event already exists
                if event_key in existing_events:
                    logger.debug(f"Skipping duplicate: {event_key[1]}")
                    continue
                
                # Add new row
                next_row = self.sheet.max_row + 1
                row_data = [
                    event.get("sport", ""),
                    event.get("event_name", ""),
                    event.get("broadcaster", ""),
                    event.get("event_date", ""),
                    event.get("location", ""),
                    event.get("teams", ""),
                    event.get("source", ""),
                    datetime.now().isoformat(),
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    cell = self.sheet.cell(row=next_row, column=col_idx)
                    cell.value = _sanitize(value)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                existing_events.add(event_key)
                new_count += 1
            
            # Save file
            if new_count > 0 or not self.file_path.exists():
                self.workbook.save(self.file_path)
                logger.info(f"Saved {new_count} new events to Excel: {self.file_path}")
            else:
                logger.info("No new events to add (all duplicates)")
            
        except Exception as e:
            logger.error(f"Failed to append events to Excel: {e}")
            raise
    
    def get_all_rows(self):
        """
        Get all rows from the Excel file
        
        Returns:
            list: List of rows (each row is a list of values)
        """
        try:
            self._load_or_create_workbook()
            
            rows = []
            for row_idx in range(1, self.sheet.max_row + 1):
                row_data = []
                for col_idx in range(1, len(self.columns) + 1):
                    cell = self.sheet.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                rows.append(row_data)
            
            return rows
        except Exception as e:
            logger.error(f"Failed to read Excel file: {e}")
            raise
    
    def clear_file(self):
        """Clear all data and recreate header"""
        try:
            self._load_or_create_workbook()
            
            # Delete all rows except header
            for row_idx in range(self.sheet.max_row, 1, -1):
                self.sheet.delete_rows(row_idx)
            
            self.workbook.save(self.file_path)
            logger.info("Excel file cleared")
        except Exception as e:
            logger.error(f"Failed to clear Excel file: {e}")
            raise
